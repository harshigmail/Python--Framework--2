import openpyxl


class HomePageData:
    # test_homePage_data = [{"firstname": "Rahul", "Email": "rahul@gmail.com", "Password": "password1"},
                        # {"firstname": "Lak", "Email": "lak@gmail.com", "Password": "password2"}]

    # or
    @staticmethod  # if this is declared without creating object we can easily call the below method
    def testData(test_case_name):  # self parameter is applied to not statis method

        data_dict = {}

        book = openpyxl.load_workbook("C:\\Users\\harshitha.n\\OneDrive\\Desktop\\Harshitha Notes\\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[data_dict]

