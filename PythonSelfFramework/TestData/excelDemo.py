import openpyxl

book = openpyxl.load_workbook("C:\\Users\\harshitha.n\\OneDrive\\Desktop\\Harshitha Notes\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)  # read value
sheet.cell(row=2, column=2).value = "Rahul"  # write into the sheet
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)  # print the maximum rows count which contains data
print(sheet.max_column)  # print the maximum columns count which contains data

# print(sheet['A2'].value) or sheet.cell(row=2, column=A or 1)

# print all the values in the excel sheet
for j in range(1, sheet.max_column + 1):
    for i in range(1, sheet.max_row + 1):
        print(sheet.cell(row=i, column=j).value)

# or

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)

# print only testcase 2 values
print("*******************************************************")
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)


# print only testcase 2 values and don't print testcase as a value in the output
print("*******************************************************")
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)