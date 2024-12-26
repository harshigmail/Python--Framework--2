import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


def update_excel_data(file_path, searchTerm, ColumnName, newValue):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == ColumnName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = newValue
    book.save(file_path)


driver = webdriver.Chrome()
file_path = "C:\\Users\\harshitha.n\\OneDrive\\Desktop\\Harshitha Notes\\download.xlsx"
fruit_name = 'Apple'
newValue = "6000"
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()

driver.find_element(By.ID, "downloadButton").click()

# edit the excel with updated value
update_excel_data(file_path, fruit_name, "price", newValue)

file_input = driver.find_element(By.CSS_SELECTOR, "input[type ='file']")
file_input.send_keys(file_path)
wait = WebDriverWait(driver, 5)
toaster_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toaster_locator))
print(driver.find_element(*toaster_locator).text)
# to get the value on the column on UI
Price_column_value = driver.find_element(By.XPATH, "//div[text() = 'Price']").get_attribute("data-column-id")
Actual_price = driver.find_element(By.XPATH, "//div[text() ='" + fruit_name + "']/parent::div/parent::div/div[@id= 'cell-" + Price_column_value + "-undefined']").text

assert Actual_price == newValue
